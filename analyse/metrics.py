from conf.config import *
import pandas as pd
from dao.db_pool import get_engine
# import openpyxl
from openpyxl import Workbook
import json
from openpyxl.styles import numbers
import numpy as np

sql = """
select 
	b.y,b.total_hldr_eqy_inc_min_int as e,
	i.n_income as r,
	i.n_income /total_hldr_eqy_exc_min_int as roe,
	d.cash_div_tax*d.base_share as s
from 
	balancesheet b,income i , fina_indicator f,dividend d
where 1=1
	and b.ts_code = i.ts_code and i.ts_code = f.ts_code and f.ts_code = d.ts_code
	and b.y = i.y and i.y = f.y and f.y	= d.y
	and b.m = i.m and i.m = f.m
	and b.ts_code = '{}' and b.y between {} and {} and b.m = 12 
	order by b.y asc;

""".format(TEST_TS_CODE_GSYH, 2006, 2010)

df = pd.read_sql_query(sql, get_engine())


# for i in df.iterrows():
#     print(i)
def my_e1():
    row_num = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    wb = Workbook()

    s = wb.active

    # _cell = s.cell('B5')
    # _cell.style.number_format.format_code = '0.00E+00'

    # sheet['A1'] = 200
    # sheet['A2'] = 300
    # sheet['A3'] = '=SUM(A1:A2)'
    # wb.save('writeFormula.xlsx')
    def r(j):
        return str(3 + j)

    i = 0
    c = row_num[i:i + 1]
    i += 1

    s[c + r(2)] = 'RP'
    s[c + r(3)] = 'NA'
    s[c + r(4)] = 'ROE'
    s[c + r(5)] = 'SOB'
    s[c + r(6)] = 'SOBr'

    # for i, row in yms.iterrows():
    #     first_trade_date_str = row['first'].strftime('%Y%m%d')

    for i, row in df.iterrows():
        c = row_num[i + 1:i + 2]

        s[c + r(1)] = row['y']
        s[c + r(2)] = row['e']
        s[c + r(3)] = row['r']
        s[c + r(4)] = '=' + c + r(3) + '/' + c + r(2)
        s[c + r(5)] = row['s']
        s[c + r(6)] = 'sobr'

        c = s.cell(row, 2, row['y'])
        c.number_format = numbers.BUILTIN_FORMATS[4]

        c.number_format = numbers.FORMAT_NUMBER

    # _cell.number_format = '#,##0.00'

    wb.save('writeFormula.xlsx')


def populate_sheet(json_data, sheet):
    sheet.cell(1, 1, "Month")
    sheet.cell(1, 2, "food")
    sheet.cell(1, 3, "heating")
    sheet.cell(1, 4, "rent")

    row = 1
    for month in json_data.keys():
        row += 1
        sheet.cell(row, 1, month)
        c = sheet.cell(row, 2, float(json_data[month]["food"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]
        c = sheet.cell(row, 3, float(json_data[month]["heating"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]
        c = sheet.cell(row, 4, float(json_data[month]["rent"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]


#############################################
#  MAIN
#############################################
if __name__ == '___main__':

    json_data = {}

    with open("original.json") as json_file:
        json_data = json.load(json_file)

    wb = Workbook()
    # When you make a new workbook you get a new blank active sheet
    # We need to delete it since we do not want it
    wb.remove(wb.active)

    for year in json_data.keys():
        sheet = wb.create_sheet(title=year)
        populate_sheet(json_data[year], sheet)

    # Save it to excel
    wb.save("formatted.xlsx")

row_num = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def my_e2():
    wb = Workbook()

    sheet = wb.active

    for i in range(5):
        c = row_num[3 - 1]

        sheet.column_dimensions[c].auto_size = True

        sheet.column_dimensions[c].bestFit = True

        sheet.column_dimensions[c].collapsed = False

    heads = 'metrics,RP,NA,ROE,SOB,SOBR'.split(',')
    i = 0;
    for head in heads:
        i += 1
        c = sheet.cell(1, i, head)

    for i, row in df.iterrows():
        y = row['y']
        e = row['e']
        r = row['r']
        s = row['s']
        c = row_num[3 - 1]

        cell = sheet.cell(i + 2, 1, y)
        cell = sheet.cell(i + 2, 2, r)
        cell.number_format = numbers.BUILTIN_FORMATS[4]
        cell = sheet.cell(i + 2, 3, e)
        cell.number_format = numbers.BUILTIN_FORMATS[4]

        cell = sheet.cell(i + 2, 4, f'=({c}{i + 2}/{c}{i + 1})-1')
        cell.number_format = numbers.FORMAT_PERCENTAGE_00
        cell = sheet.cell(i + 2, 5, s)
        cell.number_format = numbers.BUILTIN_FORMATS[4]
        cell = sheet.cell(i + 2, 6, 'sobr')

    print(sheet.max_row)

    wb.save("mye21.xlsx")


# my_e2()

def my_e3(ts_code, y, m=None, limit=5):
    # Base info
    liability = pd.read_sql_query(f"select * from liability where y = {y};", get_engine())
    aaa = liability.loc[0]['ratio']
    inflation = 2
    discount = aaa + inflation

    # Market Price Info
    if m:
        sql = f"select close,total_share from daily_basic_month where y = {y} and m = {m} and ts_code = '{ts_code}';"
    else:
        sql = f"select close,total_share from daily_basic where ts_code = '{ts_code}' ;"

    daily_info = pd.read_sql_query(sql, get_engine())

    total_share = daily_info.loc[0]['total_share']
    close = daily_info.loc[0]['close']

    # Company Net Profit Growth
    if m:
        sql = f"select y,m,netprofit_yoy from fina_indicator where ts_code = '{ts_code}' and y = {y} and m = {m} order by y desc limit {limit};"
    else:
        sql = f"select y,m,netprofit_yoy from fina_indicator where ts_code = '{ts_code}' and m = 12 order by y desc limit {limit};"

    fina = pd.read_sql_query(sql, get_engine())

    # fina['netprofit_yoy'] = fina['netprofit_yoy'].apply(lambda x: 100 if x > 100 else x)

    net_profit_growth_mean = fina['netprofit_yoy'].mean()

    sql = f"select y,n_income_attr_p/power(10,8) as profit from income where ts_code = '{ts_code}' and m = 12 order by y desc limit 1; "
    income = pd.read_sql_query(sql, get_engine())
    profit = income.loc[0]['profit']

    net_profit_growth_mean = 1 + net_profit_growth_mean / 100
    discount = 1 + discount / 100
    aaa = aaa / 100
    inflation = inflation / 100
    multi = pd.DataFrame({'growth': np.logspace(1, 5, num=5, endpoint=True, base=net_profit_growth_mean, dtype=None),
                          'discount': np.logspace(1, 5, num=5, endpoint=True, base=1 / discount, dtype=None),
                          'y': range(y - 5, y)})

    multi['profit'] = profit
    multi['future_profits'] = multi['profit'] * multi['growth'] * multi['discount']
    sustain = multi.loc[len(multi) - 2, ['future_profits']] / aaa
    multi.loc[len(multi) - 1, ['future_profits']] = sustain

    calc = multi['future_profits'].sum()

    sql = f"INSERT INTO calc_val VALUES ('{ts_code}', {aaa}, {inflation}, {y}, {'null' if m is None else m}, {multi.loc[1]['future_profits']}, {multi.loc[2]['future_profits']}, {multi.loc[3]['future_profits']}, {multi.loc[4]['future_profits']}, {calc});"
    # get_engine().execute(sql)

    print(sql)


my_e3('600516.SH', 2020)
