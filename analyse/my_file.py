from conf.config import *
import pandas as pd
from dao.db_pool import get_engine
# import openpyxl
from openpyxl import Workbook
import json
from openpyxl.styles import numbers
import numpy as np
from analyse.my_data import *


# for i in df.iterrows():
#     print(i)
def my_e1():
    row_num = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    wb = Workbook()

    s = wb.active

    # _cell = s.cell('B5')
    # _cell.style.number_format.format_code = '0.00E+00'

    # sheet['A1'] = 200
    # sheet['A2'] = 300
    # sheet['A3'] = '=SUM(A1:A2)'
    # wb.save('writeFormula.xlsx')
    def r(j):
        return str(3 + j)

    i = 0
    c = row_num[i:i + 1]
    i += 1

    s[c + r(2)] = 'RP'
    s[c + r(3)] = 'NA'
    s[c + r(4)] = 'ROE'
    s[c + r(5)] = 'SOB'
    s[c + r(6)] = 'SOBr'

    # for i, row in yms.iterrows():
    #     first_trade_date_str = row['first'].strftime('%Y%m%d')

    for i, row in df.iterrows():
        c = row_num[i + 1:i + 2]

        s[c + r(1)] = row['y']
        s[c + r(2)] = row['e']
        s[c + r(3)] = row['r']
        s[c + r(4)] = '=' + c + r(3) + '/' + c + r(2)
        s[c + r(5)] = row['s']
        s[c + r(6)] = 'sobr'

        c = s.cell(row, 2, row['y'])
        c.number_format = numbers.BUILTIN_FORMATS[4]

        c.number_format = numbers.FORMAT_NUMBER

    # _cell.number_format = '#,##0.00'

    wb.save('writeFormula.xlsx')


def populate_sheet(json_data, sheet):
    sheet.cell(1, 1, "Month")
    sheet.cell(1, 2, "food")
    sheet.cell(1, 3, "heating")
    sheet.cell(1, 4, "rent")

    row = 1
    for month in json_data.keys():
        row += 1
        sheet.cell(row, 1, month)
        c = sheet.cell(row, 2, float(json_data[month]["food"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]
        c = sheet.cell(row, 3, float(json_data[month]["heating"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]
        c = sheet.cell(row, 4, float(json_data[month]["rent"]))
        c.number_format = numbers.BUILTIN_FORMATS[4]


#############################################
#  MAIN
#############################################
if __name__ == '___main__':

    json_data = {}

    with open("original.json") as json_file:
        json_data = json.load(json_file)

    wb = Workbook()
    # When you make a new workbook you get a new blank active sheet
    # We need to delete it since we do not want it
    wb.remove(wb.active)

    for year in json_data.keys():
        sheet = wb.create_sheet(title=year)
        populate_sheet(json_data[year], sheet)

    # Save it to excel
    wb.save("formatted.xlsx")

row_num = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'


def my_e2():
    wb = Workbook()

    sheet = wb.active

    for i in range(5):
        c = row_num[3 - 1]

        sheet.column_dimensions[c].auto_size = True

        sheet.column_dimensions[c].bestFit = True

        sheet.column_dimensions[c].collapsed = False

    heads = 'metrics,RP,NA,ROE,SOB,SOBR'.split(',')
    i = 0
    for head in heads:
        i += 1
        c = sheet.cell(1, i, head)

    for i, row in df.iterrows():
        y = row['y']
        e = row['e']
        r = row['r']
        s = row['s']
        c = row_num[3 - 1]

        cell = sheet.cell(i + 2, 1, y)
        cell = sheet.cell(i + 2, 2, r)
        cell.number_format = numbers.BUILTIN_FORMATS[4]
        cell = sheet.cell(i + 2, 3, e)
        cell.number_format = numbers.BUILTIN_FORMATS[4]

        cell = sheet.cell(i + 2, 4, f'=({c}{i + 2}/{c}{i + 1})-1')
        cell.number_format = numbers.FORMAT_PERCENTAGE_00
        cell = sheet.cell(i + 2, 5, s)
        cell.number_format = numbers.BUILTIN_FORMATS[4]
        cell = sheet.cell(i + 2, 6, 'sobr')

    print(sheet.max_row)

    wb.save("mye21.xlsx")


# my_e2()

def my_e3(ts_code, y, m=None, limit=5):
    # Base info
    liability = pd.read_sql_query(f"select * from liability where y = {y};", get_engine())
    aaa = liability.loc[0]['ratio']
    inflation = 2
    discount = aaa + inflation

    # Market Price Info
    if m:
        sql = f"select close,total_share from daily_basic_month where y = {y} and m = {m} and ts_code = '{ts_code}';"
    else:
        sql = f"select close,total_share from daily_basic where ts_code = '{ts_code}' ;"

    daily_info = pd.read_sql_query(sql, get_engine())

    total_share = daily_info.loc[0]['total_share']
    close = daily_info.loc[0]['close']

    # Company Net Profit Growth
    if m:
        sql = f"select y,m,netprofit_yoy from fina_indicator where ts_code = '{ts_code}' and y = {y} and m = {m} order by y desc limit {limit};"
    else:
        sql = f"select y,m,netprofit_yoy from fina_indicator where ts_code = '{ts_code}' and m = 12 order by y desc limit {limit};"

    fina = pd.read_sql_query(sql, get_engine())

    # fina['netprofit_yoy'] = fina['netprofit_yoy'].apply(lambda x: 100 if x > 100 else x)

    net_profit_growth_mean = fina['netprofit_yoy'].mean()

    sql = f"select y,n_income_attr_p/power(10,8) as profit from income where ts_code = '{ts_code}' and m = 12 order by y desc limit 1; "
    income = pd.read_sql_query(sql, get_engine())
    profit = income.loc[0]['profit']

    net_profit_growth_mean = 1 + net_profit_growth_mean / 100
    discount = 1 + discount / 100
    aaa = aaa / 100
    inflation = inflation / 100
    multi = pd.DataFrame({'growth': np.logspace(1, 5, num=5, endpoint=True, base=net_profit_growth_mean, dtype=None),
                          'discount': np.logspace(1, 5, num=5, endpoint=True, base=1 / discount, dtype=None),
                          'y': range(y - 5, y)})

    multi['profit'] = profit
    multi['future_profits'] = multi['profit'] * multi['growth'] * multi['discount']
    sustain = multi.loc[len(multi) - 2, ['future_profits']] / aaa
    multi.loc[len(multi) - 1, ['future_profits']] = sustain

    calc = multi['future_profits'].sum()

    sql = f"INSERT INTO calc_val VALUES ('{ts_code}', {aaa}, {inflation}, {y}, {'null' if m is None else m}, {multi.loc[1]['future_profits']}, {multi.loc[2]['future_profits']}, {multi.loc[3]['future_profits']}, {multi.loc[4]['future_profits']}, {calc});"
    # get_engine().execute(sql)

    print(sql)


def tangchao(ts_code, y):
    aaa = pd.read_sql_query(f"select * from liability  where y = {y};", get_engine()).iloc[0, 2]
    aaa = aaa / 100
    x2aaa = aaa * 2
    expect_ep = 1 / x2aaa

    daily = pd.read_sql_query(f"select ts_code,close,total_share from daily_basic where ts_code = '{ts_code}';",
                              get_engine())
    # price = daily.iloc[0, 1]
    total_share = daily.iloc[0, 2] / 10000
    # market = price * total_share
    """
    growth
    """
    df = pd.read_sql_query(f"select 1 from fina_indicator where ts_code = '{ts_code}' and y = {y} and m = 12;",
                           get_engine())

    if len(df):
        roes = pd.read_sql_query(
            f"select roe from fina_indicator where ts_code = '{ts_code}' and y between {y} - 3 and {y};",
            get_engine())
        rase_mean = roes.mean()
        incomes = pd.read_sql_query(
            f"select y,m,n_income/100000000 as n_income from income where ts_code= '{ts_code}' and y = {y}",
            get_engine())

        retained_profits = incomes.iloc[0]['n_income']

    else:
        roes = pd.read_sql_query(f"select y,m,roe from fina_indicator where ts_code = '{ts_code}' and y > {y} - 4;",
                                 get_engine())
        # get session’s month report number
        session_month_num = roes[roes['y'] == y]['m'].max()

        all_before_session_roes = roes[(roes['y'] != y) & (roes['m'] == session_month_num)]
        all_before_year_end_roes = roes[(roes['y'] != y) & (roes['m'] == 12)]
        # before join , need make the index accordance
        all_before_session_roes = all_before_session_roes.reset_index()
        all_before_year_end_roes = all_before_year_end_roes.reset_index()

        # infer the nearest roe by before session
        before_roes_2_month_ratio = (all_before_year_end_roes['roe'] / all_before_session_roes['roe']).mean()

        # all before roe number series
        all_roes_to_mean = all_before_year_end_roes['roe']

        # use ratio and now month's roe calc year end's roe
        # replace the farthest year's roe ,to prepare calc de avg roe
        all_roes_to_mean[0] = roes[(roes['y'] == y) & (roes['m'] == session_month_num)].iloc[0][
                                  'roe'] * before_roes_2_month_ratio
        rase_mean = all_roes_to_mean.mean()

        """
            profit ttm
        """
        incomes = pd.read_sql_query(
            f"select y,m,n_income/100000000 as n_income from income where ts_code= '{ts_code}' and y >= {y} -1",
            get_engine())
        last_y_m = incomes[(incomes['y'] == y - 1) & (incomes['m'] == session_month_num)].iloc[0]['n_income']
        last_y_12 = incomes[(incomes['y'] == y - 1) & (incomes['m'] == 12)].iloc[0]['n_income']
        y_m = incomes[(incomes['y'] == y) & (incomes['m'] == session_month_num)].iloc[0]['n_income']
        retained_profits = last_y_12 - last_y_m + y_m

    # 折现
    discounting = 1 / 1.06
    # 增长
    rase_mean_rate = 1 + rase_mean / 100

    # 两段折现基础数据
    d_arr = np.logspace(1, 4, 4, endpoint=True, base=discounting)
    i_arr = np.logspace(1, 4, 4, endpoint=True, base=rase_mean_rate)
    # 第一段
    real_increase_rate = d_arr * i_arr
    # 永续段替换
    real_increase_rate[3] = d_arr[3] / aaa * real_increase_rate[2]

    sum_all_multi = real_increase_rate.sum()
    calc_market_value = sum_all_multi * retained_profits
    # print(price)
    print(f"calc_market_value:{calc_market_value}")
    print("calc_market_price:{}".format(calc_market_value / total_share))
    print("calc_market_price_count:{}".format(calc_market_value * 0.7 / total_share))


def get_calc_param():
    param = {}
    # param['aaa'] = get_liability()
    param['aaa'] = 0.05
    param['premium'] = 0.03
    param['inflation'] = 0
    param['discount'] = param['aaa'] + param['premium'] + param['inflation']
    param['enternity'] = 0.05
    param['calc_discount'] = 0.8
    param['safety_margin'] = 0.7
    return param


def my_calc(ts_code, param):
    discount = param['discount']
    enternity = param['enternity']
    calc_discount = param['calc_discount']
    # get total share
    daily = pd.read_sql_query(
        f"select close,total_share,total_mv,pe,pe_ttm from daily_basic where ts_code = '{ts_code}';",
        get_engine())

    total_share = daily.iloc[0, 1] / 10000
    total_mv = daily.iloc[0, 2] / 10000
    pe = daily.iloc[0, 3]
    pe_ttm = daily.iloc[0, 4]

    # growth
    growth_df = pd.read_sql_query(
        f"select avg(netprofit_yoy) as growth ,avg(roe) as as_pe from i_data where ts_code = '{ts_code}';",
        get_engine())

    growth = growth_df.iloc[0, 0] / 100
    growth *= calc_discount

    # income
    y = datetime.datetime.now().year
    income = pd.read_sql_query(
        f"select y,m,n_income/100000000 as n_income from income where ts_code= '{ts_code}' and y = {y} -1 and m = 12 ",
        get_engine())

    income_0 = income.iloc[0, 2]
    income_1 = income_0 * (1 + growth) ** 1 / (1 + discount) ** 1
    income_2 = income_0 * (1 + growth) ** 2 / (1 + discount) ** 2
    income_3 = income_0 * (1 + growth) ** 3 / (1 + discount) ** 3

    # Cash Flow Discount Calc
    forever = income_3 * (1 + enternity) / (discount - enternity)
    estimate_value = income_1 + income_2 + income_3 + forever

    chipe_rate = total_mv / estimate_value

    estimate_value_safety_margin = estimate_value * param['safety_margin']

    # Tangchao Eazy Calc
    as_pe = growth_df.iloc[0, 1]
    tangchao_calc = income_3 * as_pe
    # tangchao_calc = tangchao_calc * param['safety_margin']

    if total_mv < tangchao_calc * 0.5:
        tangchao_flag = 1
    elif total_mv < tangchao_calc * 0.8333:
        tangchao_flag = 2
    elif total_mv < tangchao_calc * 1.3333:
        tangchao_flag = 3
    elif total_mv < tangchao_calc * 1.6666:
        tangchao_flag = 4
    else:
        tangchao_flag = 5

    result = locals()
    del result['param']
    del result['daily']
    del result['income']
    del result['growth_df']

    print(result)
    return result


def loop_codes():
    param = get_calc_param()
    # codes = pd.read_sql_query("select DISTINCT ts_code from i_data where debt_to_assets < 50 ts_code = '600177.SH' ;", get_engine())
    codes = pd.read_sql_query("""select DISTINCT ts_code from i_data   ;""", get_engine())
    mps = []
    for i, row in codes.iterrows():
        t = my_calc(row['ts_code'], param)
        mps.append(t)
    df = pd.DataFrame(mps)
    df.to_sql('i_calc', get_engine(), index=False, if_exists='replace')


if __name__ == '__main__':
    #
    # r = my_calc(TEST_TS_CODE_GZMT, param)
    # from collections import OrderedDict
    #
    # r = OrderedDict(sorted(r.items(), key=lambda t: t[0], reverse=True))
    # print(r)
    loop_codes()
