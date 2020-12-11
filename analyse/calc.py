import conf.config as config
import dao.db_dao as dao
import pandas as pd
from time import time
from dao.db_pool import get_engine
from conf.config import *
# import matplotlib
# matplotlib.use('Agg')
# from matplotlib import pyplot as mp
from dao.tushare_oop_dao import initOne
import matplotlib.pyplot as mp


def test_calc_repay(ts_code, start, end, begin_position=100):
    df = dao.get_dividend(ts_code, start, end)

    inc_cash = 0.0
    inc_stock = begin_position

    df = df[::-1]
    # for row in df.itertuples():
    for i, row in df.iterrows():
        # for row in df.items():
        cash = row['cash_div']
        stk = row['stk_div']
        if cash:
            inc_cash += inc_stock * cash
        if stk:
            inc_stock *= 1 + stk
        # print("i:{}\tstk:{}\tcash:{}\ti_cash:{}\ti_stk:{}".format(i, stk, cash, inc_cash, inc_stock))
        print("{}\t{}\t{}\t{}\t{}".format(i, stk, cash, inc_cash, inc_stock))
    return inc_stock, inc_cash


def test_one_repay(ts_code, start_year, period):
    end_year = period + start_year
    first_begin, first_end = dao.get_trade_date(start_year, 1)
    last = dao.get_trade_date(end_year, 1)[1]
    df_price_1 = dao.get_stock_price_monthly(ts_code, first_end)
    df_price_2 = dao.get_stock_price_monthly(ts_code, last)
    p_open = df_price_1.loc[0]['open']
    p_close = df_price_2.loc[0]['close']

    position_begin = 100

    position, cash = test_calc_repay(ts_code, first_end, last, begin_position=position_begin)

    principal = position_begin * p_open
    value = position * p_close + cash

    earnings = value - principal
    earnings_percent = earnings / principal

    print('first_end:\t', first_end)
    print('last:\t', last)
    print('p_open:\t', p_open)
    print('p_close:\t', p_close)
    print('position:\t', position)
    print('cash:\t', cash)
    print('principal:\t', principal)
    print('value:\t', value)
    print('earnings:\t', earnings)
    print('earnings_percent:\t', earnings_percent * 100)


class Analyser:
    def __init__(self, y, m, earning_duration=None, earning_mean_year=None, earning_inc_ratio=None,
                 underrate_ep_2_3A_multi=None,
                 underrate_divident_2_3A=None, low_percent=None, underrate_price_2_touch_assert=None,
                 underrate_price_2_cur_ass_val=None,
                 ):
        self.__y = y
        self.__m = m
        self.earning_duration = 5 if earning_duration is None else earning_duration
        self.earning_mean_year = 3 if earning_mean_year is None else earning_mean_year
        self.earning_inc_ratio = 1.07 if earning_inc_ratio is None else earning_inc_ratio

        # 1. ep L3A*2
        self.underrate_ep_2_3A_multi = 2 if underrate_ep_2_3A_multi is None else underrate_ep_2_3A_multi
        # 2. dividend > L3A 2/3
        self.underrate_divident_2_3A = 2 / 3 if underrate_divident_2_3A is None else underrate_divident_2_3A
        # 3. pe in lowest 10%
        self.low_percent = 1 if low_percent is None else low_percent
        # 4. price < touch assert 2/3
        self.underrate_price_2_touch_assert = 2 / 3 if underrate_price_2_touch_assert is None else underrate_price_2_touch_assert
        # 5. price < total_cur_assets VALUE 2/3
        self.underrate_price_2_cur_ass_val = 2 / 3 if underrate_price_2_cur_ass_val is None else underrate_price_2_cur_ass_val

        self.stat = {}

        self.__result = None

        self.standard_table_name = 'standard'
        self.standard_start_table_name = 'standard_stat'

    def process(self):
        if self.__load_db():
            print('{} {} load db matrix ok'.format(self.__y, self.__m))
            return
        print('{} {} analyse start'.format(self.__y, self.__m))
        self.__metadata()
        self.exec_mask()

    def __load_db(self):
        stat = dao.get_standard_stat(self.__y, self.__m)
        if stat is not None and len(stat) > 0:
            self.stat = stat
            self.__result = dao.get_standard(self.__y, self.__m)
            return 'ok'
        return None

    def __metadata(self):
        start = time()

        a3 = dao.get_liability(self.__y)

        pe_threshold = dao.get_pe_low(self.__y, self.__m, self.low_percent)
        start_year = self.__y - self.earning_duration - self.earning_mean_year

        df = dao.get_analyse_collection(self.__y, self.__m, pe_threshold, start_year)

        print('=' * 32, 'load cost:', time() - start)
        start = time()
        """
            UNDERRATE
        """

        # 1. ep L3A*2
        self.stat['ep_2_3A'] = df['ep'] * 100 > a3 * self.underrate_ep_2_3A_multi
        # 2. dividend > L3A 2/3
        self.stat['dividend_2_3A'] = df['dv_ratio'] > a3 * self.underrate_divident_2_3A

        # 3. pe in lowest 10%
        # add above

        # 4. price < tangible_asset 2/3
        # self.stat['price_touch_assert'] = (df['total_mv'] * 10000 / self.underrate_price_2_touch_assert) < df['total_assets'] - df['intan_assets'] - df['r_and_d'] - df['goodwill']
        # 'intan_assets'] - df['r_and_d'] - df['goodwill'] - df['lt_amor_exp'] - df['defer_tax_assets']

        self.stat['price_touch_assert_exc_1'] = (df['total_mv'] * 10000) / 2 / 3 < df['tangible_asset']
        # self.stat['price_touch_assert_exc_066'] = (df['total_mv'] * 10000 / 0.666 ) < df['total_hldr_eqy_exc_min_int']

        # 5. price < total_cur_assets VALUE 2/3
        # self.stat['price_2_cur_asset_val'] = df['total_mv'] * 10000 / self.underrate_price_2_cur_ass_val < df['total_cur_assets'] - df['total_liab']

        """
            FINANCIAL
        """
        # 1, current_ratio > 2 |  quick_ratio  > 1
        self.stat['cur_quick_ratio'] = (df['current_ratio'] > 2) | (df['quick_ratio'] > 1)

        # 2. debt-to-equity < 1
        # self.stat['debt-to-equity_exc'] = df['total_liab'] < df['total_hldr_eqy_exc_min_int']
        self.stat['debt-to-equity_inc'] = df['total_liab'] < df['total_hldr_eqy_inc_min_int']

        # Current Assets VAL > debt 1/2
        self.stat['cur_asset_val_2_debt'] = df['total_cur_assets'] - df['total_liab'] > df['total_liab'] / 2

        """
        EARNING
        """
        # 1. 10 year earning rate 7%
        # 2. 10 year earning rate de decrease more than 5% below 2year
        print('=' * 32, 'analyse cost:', time() - start)
        start = time()

        detail_list = dao.get_fina(df.ts_code, self.__y - self.earning_duration - self.earning_mean_year, self.__y,
                                   self.__m)
        print('=' * 32, 'load cost:', time() - start)
        start = time()
        self.stat['earning_power'] = df['ts_code'].apply(
            lambda x: check_earning_power(x, self.__y, self.earning_duration, self.earning_mean_year, self.__m,
                                          self.earning_inc_ratio,
                                          data=detail_list))

        print('=' * 32, 'analyse cost:', time() - start)
        # start = time()

        self.__result = df

    def exec_mask(self):
        df = self.__result
        # total count
        total = len(df)
        r = None
        for k, mask in self.stat.items():
            self.stat[k] = mask.sum() / total
            if r is None:
                r = mask
            else:
                r = r & mask
            print('{}\t:{}'.format(k, r.sum()))
        df = df[r]

        self.stat['total'] = total
        self.stat['all'] = len(df)
        self.stat['all_ratio'] = len(df) / total

        df_stat = pd.DataFrame(self.stat, index=pd.Series([self.__y]))
        df_stat.insert(0, 'm', self.__m)
        df_stat.insert(0, 'y', self.__y)

        df.to_sql(self.standard_table_name, get_engine(), index=False, if_exists='append')
        df_stat.to_sql(self.standard_start_table_name, get_engine(), index=False, if_exists='append')

        self.stat = df_stat
        self.__result = df


def analyse_pe():
    start = 1994
    y = []
    p_10 = []
    p_50 = []
    for i in range(27):
        print(i)
        year = start + i
        y.append(year)

        pe = dao.get_pe_low(year, 12, 0.1)
        p_10.append(pe)

        # pe = dao.get_pe_low(year, 12, 0.5)
        # p_50.append(pe)

    mp.plot(y, p_10, label='10%')
    # mp.plot(y, p_50,label='50%')

    mp.legend()
    mp.show()


def check_earning_power(ts_code, y, earning_duration, earning_mean_year, m, ratio, data=None):
    # print('=' * 32, 'checking', ts_code)
    if data is None:
        df = dao.get_fina(ts_code, y - earning_duration - earning_mean_year, y, m)
    else:
        df = data[data['ts_code'] == ts_code]

    if not len(df) or len(df) < earning_duration + earning_mean_year:
        return False

    try:
        a = df['eps'] > 0
    except KeyError as e:
        print(e)
        exit(4)
    if a.sum() < len(a):
        return False

    e_n = df[:earning_mean_year]['eps'].mean()
    e_f = df[:-1 - earning_mean_year:-1]['eps'].mean()
    r = e_n / e_f
    if r < ratio:
        # print('check earning power exit by ratio - > ts_code:', ts_code)
        # print(df[['y', 'eps']])
        # print(e_n)
        # print(e_f)
        return False

    a_a = df[:earning_duration]
    a_b = df[1:earning_duration + 1]
    a_a = a_a.reset_index(drop=True)
    a_b = a_b.reset_index(drop=True)
    a_c = (a_a['eps'] < a_b['eps'] * 0.95)
    if a_c.sum() > 2:
        # print('check earning power exit by flacturat- > ts_code:', ts_code)
        # print('a:', a_a)
        # print('b:', a_b)
        # print('c:', a_c)
        return False
    return True


def analys_array():
    start, m = 2000, 12
    for i in range(0, 21):
        y = start + i
        if y == 2020:
            m = 6
        a = Analyser(y, m)
        a.process()


def show():
    df = dao.get_stat()
    filters = ['all', 'all_ratio', 'total', 'y', 'm']
    for key in df.columns.values:
        if key not in filters:
            mp.plot(df[key], label=key)

    mp.legend()
    mp.show()


def show_2():
    mp.rcParams['font.sans-serif'] = ['KaiTi']
    mp.rcParams['axes.unicode_minus'] = False
    # target_list = ['洋河股份', '贵州茅台', '分众传媒', '古井贡酒', '海康威视', '科大讯飞','招商银行','中国平安']
    target_list = ['洋河股份', '贵州茅台', '分众传媒', '古井贡酒', '海康威视', '科大讯飞', '招商银行', '中国平安']

    sql = "select roe,y from fina_indicator where ts_code = '{}' and m = 12 order by end_date desc limit 10;"
    for name in target_list:
        code = dao.get_code_by_name(name)
        initOne(code)
        df = dao.get_df(sql.format(code), 'y')
        name = dao.get_name_by_code(code)
        mp.plot(df, label=name)
    mp.legend()
    mp.show()


class IncrementCalculator:
    def __init__(self, ts_code, start, end):
        self.start = start
        self.end = end
        self.ts_code = ts_code
        self._market = None
        self._premium = None
        self._capital = 10000
        self.final = None
        self.discount = 0.05
        self.results = None

    def load(self):
        # load market
        load_market = f"""select ts_code,y,m,close from daily_basic_month 
                            where ts_code = '{self.ts_code}' and {self.end} >= y and y>= {self.start}  
                                    and m = 6 order by y ,m;"""
        load_premium = f"""select * from dividend_stat where ts_code = '{self.ts_code}' and {self.end} >= y and y>= {self.start}  """;

        market_df = pd.read_sql_query(load_market, get_engine())
        premium_df = pd.read_sql_query(load_premium, get_engine())

        if len(market_df) < (self.end - self.start + 1):
            raise LookupError('Not A Good Data Sample.')

        premium_dict = {}
        for i, row in premium_df.iterrows():
            premium_dict[row['y']] = row
        self._market = market_df
        self._premium = premium_dict

    def calc(self):
        ss_end = None
        results = []
        self.results = results
        for i, market in self._market.iterrows():
            close = market['close']

            if ss_end is None:
                ss_end = self._capital / close

            ss_start = ss_end

            y = market['y']
            m = market['m']
            print('calc ing', y, m)

            if m == 6 and y in self._premium:
                div = self._premium[y]
                stk_div = div['stk_div']
                cash_div = div['cash_div']
                got_cash = ss_start * cash_div
                got_cash2stk = got_cash / close
                got_stk = ss_start * stk_div
                ss_end = ss_start + got_cash2stk + got_stk
            else:
                stk_div = 0
                cash_div = 0
                got_cash = 0
                got_cash2stk = 0
                got_stk = 0
            mv = ss_end * close
            yield_rate = (mv - self._capital) / self._capital
            yield_dicount = yield_rate / (1 + self.discount) ** (i + 1)

            stk_div = 0 if not stk_div else stk_div
            cash_div = 0 if not cash_div else cash_div
            got_cash = 0 if not got_cash else got_cash
            got_cash2stk = 0 if not got_cash2stk else got_cash2stk
            got_stk = 0 if not got_stk else got_stk

            result = {'y': y, 'stk_div': stk_div, 'cash_div': cash_div, 'close': close, 'ss_start': ss_start,
                      'got_cash': got_cash, 'got_cash2stk': got_cash2stk, 'got_stk': got_stk, 'ss_end': ss_end,
                      'mv': mv, 'yield_rate': yield_rate, 'yield_discount': yield_dicount}
            self.final = yield_dicount
            results.append(result)

    def show(self):
        self.plot()

    def plot(self):
        import matplotlib.pyplot as plt
        from matplotlib.gridspec import GridSpec

        df = pd.DataFrame(self.results)
        df = df.set_index(['y'])

        fig = plt.figure(constrained_layout=True, figsize=(12, 6))
        gs = GridSpec(3, 3, figure=fig)
        down = fig.add_subplot(gs[:2, :])
        # sub = up.twinx()  # instantiate a second axes that shares the same x-axis
        down1 = fig.add_subplot(gs[2, :])
        # df[['yield_rate']].plot(kind='semilogy', ax=down)
        # k = down.semilogy(df.index, df['yield_rate'])
        # rects = df[['yield_rate']].plot(kind='bar', ax=down)
        df['yield_rate'] = round(df['yield_rate'], 4)
        df['color'] = df['yield_rate'].apply(lambda x: 'r' if x < 0 else 'b')
        rects = down.bar(df.index, df['yield_rate'], 0.5, color=df['color'].tolist(), label='yield_rate')

        def autolabel(rects, ax):
            for rect in rects:
                height = rect.get_height()
                ax.annotate('{}'.format(height),
                            xy=(rect.get_x() + rect.get_width() / 2, height),
                            xytext=(0, 3),  # 3 points vertical offset
                            textcoords="offset points",
                            ha='center', va='bottom')

        autolabel(rects, down)
        # df[['mv']].plot(ax=sub)
        plt.show()

    def process(self):
        self.load()
        self.calc()
        # self.show()
        return self


def mult_calc(index_code, start, end, year_num):
    ics = []
    stat_years = {}
    for y in range(start, end):
        print(y)
        sql = f"""select * from index_weight where index_code = '000016.SH' and y = {y - 1} and m = 12
and con_code in (
    select ts_code from fina_indicator f where y = {y - 1} and m = 12 and roe > 15 and f.debt_to_assets < 50
)"""
        sql = f"""select distinct ts_code as con_code from i_data i where i.debt_to_assets < 50 and pe_ttm < aaa_pe/2 and i.y = {y};"""

        sql = f"""
one_fina
"""

        df = pd.read_sql_query(sql, get_engine())

        stat = {'sum': 0, 'count': 0}
        stat_years[y] = stat
        for i, row in df.iterrows():
            try:
                ic = IncrementCalculator(row['con_code'], y, y + year_num).process()
            except LookupError as m_e:
                print('LookupError' + row['con_code'])
            else:
                ics.append(ic)
                stat['sum'] = stat['sum'] + ic.final
                stat['count'] = stat['count'] + 1
        if stat['count'] and stat['sum']:
            stat['avg'] = stat['sum'] / stat['count']
        else:
            stat['avg'] = 0
    from openpyxl import Workbook
    from openpyxl.styles import numbers

    from openpyxl.styles import Alignment
    from openpyxl.styles import PatternFill, colors

    red = PatternFill("solid", fgColor=colors.RED)
    green = PatternFill("solid", fgColor=colors.GREEN)

    wb = Workbook()
    wb.remove(wb.active)
    sheet = wb.create_sheet(title='{} {}-{}'.format(ts_code, start, end))

    sheet.cell(1, 1, "ts_code")
    sheet.cell(1, 2, "start")
    sheet.cell(1, 3, "end")
    for i in range(year_num):
        sheet.cell(1, 4 + i, i + 1)

    col_list = "y,stk_div,cash_div,close,ss_start,ss_end,mv,yield_rate,yield_discount".split(',')

    row = 2
    row_span = len(col_list)
    for ic in ics:
        cell = sheet.cell(row, 1, ic.ts_code)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row + row_span - 1, end_column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = sheet.cell(row, 2, ic.start)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row + row_span - 1, end_column=2)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell = sheet.cell(row, 3, ic.end)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row + row_span - 1, end_column=3)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        for y_d in range(len(ic.results)):
            result = ic.results[y_d]
            i = 0
            for column in col_list:
                # val = float(result[column])
                val = result[column]
                c = sheet.cell(row + i, 4 + y_d, val)
                if type(val) is float:
                    c.number_format = numbers.BUILTIN_FORMATS[4]

                if column == 'yield_rate':
                    if val < 0:
                        c.fill = red
                    else:
                        c.fill = green
                i += 1

        row += row_span

    wb.save("formatted.xlsx")
    for key, val in stat_years.items():
        print(key, val)


if __name__ == '__main__':
    # show()
    # show_2()

    ts_code = TEST_TS_CODE_GZMT
    start = 2010
    end = 2018

    mult_calc(ts_code, start, end, 3)
